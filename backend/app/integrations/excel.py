"""Excel bulk-upload parsing for institution candidate loads."""

import csv
import io
from typing import Any

from openpyxl import load_workbook

# Expected column headers (case-insensitive) -> candidate field
COLUMN_MAP = {
    "name": "full_name",
    "full name": "full_name",
    "student name": "full_name",
    "phone": "phone",
    "mobile": "phone",
    "mobile number": "phone",
    "email": "email",
    "gender": "gender",
    "city": "city",
    "district": "city",
    "state": "state",
    "pincode": "pincode",
    "trade": "primary_trade",
    "primary trade": "primary_trade",
    "skill": "primary_trade",
    "course / trade / specialization": "primary_trade",
    "course / trade": "primary_trade",
    "course": "primary_trade",
    "specialization": "primary_trade",
    "experience": "experience_years",
    "experience years": "experience_years",
    "experience (months/years)": "experience_years",
    "education": "education_level",
    "qualification": "education_level",
    "certification": "certification",
    "languages": "languages",
    "expected salary": "expected_salary",
}

def _clean_header(name: str) -> str:
    """Normalize a CSV header: lower, strip, collapse spaces, remove common non-alphanumerics."""
    text = "".join(ch for ch in name.lower().strip() if ch.isalnum() or ch.isspace() or ch == "/")
    return "".join(text.split()).replace("/", "")


_INSTITUTION_HEADER_ALIASES = {
    "studentname": "full_name",
    "name": "full_name",
    "fullname": "full_name",
    "full name": "full_name",
    "mobilenumber": "phone",
    "mobile": "phone",
    "phone": "phone",
    "gender": "gender",
    "dateofbirthage": "date_of_birth",
    "dateofbirth": "date_of_birth",
    "age": "date_of_birth",
    "qualification": "education_level",
    "education": "education_level",
    "coursetradespecialization": "primary_trade",
    "coursetrade": "primary_trade",
    "course": "primary_trade",
    "trade": "primary_trade",
    "specialization": "primary_trade",
    "primarytrade": "primary_trade",
    "skill": "primary_trade",
    "passingyearexpectedpassingyear": "passing_year",
    "passingyear": "passing_year",
    "expectedpassingyear": "passing_year",
    "currentstatuscurrentstuddentalumni": "current_status",
    "currentstatus": "current_status",
    "preferredjobrole": "preferred_job_role",
    "jobrole": "preferred_job_role",
    "district": "city",
    "city": "city",
    "state": "state",
    "willingtorelocateyesno": "willing_to_relocate",
    "willingtorelocate": "willing_to_relocate",
    "fresherexperienced": "experience_level",
    "experiencelevel": "experience_level",
    "experiencemonthsyears": "experience_years",
    "experience": "experience_years",
    "experienceyears": "experience_years",
    "remarksspecialskills": "remarks",
    "remarks": "remarks",
    "specialskills": "remarks",
    "institutionname": "institution_name",
    "institution": "institution_name",
    "college": "institution_name",
    "pincode": "pincode",
    "certification": "certification",
    "languages": "languages",
    "expectedsalary": "expected_salary",
}

INSTITUTION_TEMPLATE_COLUMNS = {
    **COLUMN_MAP,
    **_INSTITUTION_HEADER_ALIASES,
}

CAMPAIGN_HEADERS = [
    "name",
    "phone",
    *(header for number in range(1, 6) for header in (f"question {number}", f"answer {number}")),
]


def parse_candidate_workbook(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """Parse an .xlsx file into candidate dicts. Returns (rows, errors)."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    errors: list[str] = []
    rows: list[dict[str, Any]] = []

    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_cells:
        return [], ["Worksheet is empty."]

    headers = [_clean_header(str(h)) if h is not None else "" for h in header_cells]

    for idx, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if raw is None or all(c is None for c in raw):
            continue
        record: dict[str, Any] = {}
        for col_idx, value in enumerate(raw):
            if col_idx >= len(headers):
                break
            field = INSTITUTION_TEMPLATE_COLUMNS.get(headers[col_idx])
            if field and value is not None:
                record[field] = value

        if not record.get("full_name") or not record.get("phone") or not record.get("primary_trade"):
            errors.append(
                f"Row {idx}: missing required Student Name / Mobile Number / Trade - skipped."
            )
            continue

        record["phone"] = "".join(ch for ch in str(record["phone"]) if ch.isdigit())
        record["full_name"] = str(record["full_name"]).strip()
        for int_field in ("experience_years", "expected_salary"):
            if int_field in record:
                try:
                    record[int_field] = int(float(record[int_field]))
                except (ValueError, TypeError):
                    record.pop(int_field)
        rows.append(record)

    return rows, errors


def parse_institution_csv(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """Parse the institution placement-officer CSV template."""
    errors: list[str] = []
    rows: list[dict[str, Any]] = []
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return [], ["CSV header row is missing."]

    fieldnames = [_clean_header(name) for name in reader.fieldnames if name]
    required = {
        "studentname",
        "mobilenumber",
        "coursetradespecialization",
    }
    if not required.issubset(set(fieldnames)):
        missing = ", ".join(required - set(fieldnames))
        return [], [f"CSV is missing required columns: {missing}. Detected headers: {fieldnames}"]

    header_map = {_clean_header(key): key for key in reader.fieldnames if key}
    for idx, raw in enumerate(reader, start=2):
        record: dict[str, Any] = {}
        for key, value in raw.items():
            if key is None:
                continue
            clean_key = _clean_header(key)
            field = INSTITUTION_TEMPLATE_COLUMNS.get(clean_key)
            if field and value is not None and str(value).strip() != "":
                record[field] = str(value).strip()
        if not record.get("full_name") or not record.get("phone") or not record.get("primary_trade"):
            errors.append(
                f"Row {idx}: missing required Student Name / Mobile Number / Trade - "
                f"found full_name={bool(record.get('full_name'))} phone={bool(record.get('phone'))} "
                f"trade={bool(record.get('primary_trade'))}."
            )
            continue
        record["phone"] = "".join(ch for ch in record["phone"] if ch.isdigit())
        for int_field in ("experience_years", "expected_salary"):
            if int_field in record:
                try:
                    record[int_field] = int(float(record[int_field]))
                except (ValueError, TypeError):
                    record.pop(int_field)
        rows.append(record)
    return rows, errors


def _normalize_campaign_header(name: str) -> str:
    """Normalize campaign headers: lower case, strip spaces, collapse non-alphanumerics."""
    text = "".join(ch for ch in name.lower().strip() if ch.isalnum() or ch.isspace())
    return "".join(text.split())


def parse_campaign_workbook(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """Parse the fixed five-question Facebook campaign workbook template."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    errors: list[str] = []
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_cells:
        return [], ["Worksheet is empty."]

    headers = [str(header).strip() if header is not None else "" for header in header_cells]
    normalized_headers = [_normalize_campaign_header(h) for h in headers]
    
    # Map header positions based on normalized keys or aliases
    header_positions = {}
    for index, nh in enumerate(normalized_headers):
        if not nh:
            continue
        # Check name and phone aliases
        if nh in ("name", "fullname", "studentname", "candidatename"):
            header_positions["name"] = index
        elif nh in ("phone", "mobile", "mobilenumber", "phonenumber", "phone_number"):
            header_positions["phone"] = index
        else:
            # Check question/answer patterns (e.g. question1, q1, answer1, a1)
            for number in range(1, 6):
                if nh in (f"question{number}", f"q{number}"):
                    header_positions[f"question {number}"] = index
                elif nh in (f"answer{number}", f"a{number}"):
                    header_positions[f"answer {number}"] = index

    # Validate missing headers
    missing_headers = []
    if "name" not in header_positions:
        missing_headers.append("Name")
    if "phone" not in header_positions:
        missing_headers.append("Phone")
    for number in range(1, 6):
        if f"question {number}" not in header_positions:
            missing_headers.append(f"Question {number}")
        if f"answer {number}" not in header_positions:
            missing_headers.append(f"Answer {number}")

    if missing_headers:
        return [], [f"Missing required columns: {', '.join(missing_headers)}."]

    rows: list[dict[str, Any]] = []
    for row_number, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not raw or all(value is None for value in raw):
            continue

        def cell(header: str) -> str | None:
            pos = header_positions.get(header)
            if pos is not None and pos < len(raw):
                value = raw[pos]
                return str(value).strip() if value is not None else None
            return None

        full_name = cell("name")
        phone = cell("phone")
        if not full_name or not phone:
            errors.append(f"Row {row_number}: missing required Name or Phone - skipped.")
            continue
            
        cleaned_phone = "".join(ch for ch in phone if ch.isdigit())

        rows.append(
            {
                "full_name": full_name,
                "phone": cleaned_phone,
                "responses": [
                    {
                        "question_number": number,
                        "question": cell(f"question {number}"),
                        "answer": cell(f"answer {number}"),
                    }
                    for number in range(1, 6)
                ],
            }
        )
    return rows, errors
